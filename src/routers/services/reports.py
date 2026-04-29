from sqlalchemy.orm import Session
from database.models import *
from utils.formatters import format_group_name
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import date
from copy import copy
import io


def generate_social_passport(group_id: int, db: Session) -> tuple:
    group = db.query(Group).filter_by(group_id=group_id).first()
    if not group:
        return None

    group_name = format_group_name(group)

    students_in_group = db.query(StudentInGroup).filter_by(group_id=group_id).all()
    student_ids = [s.student_id for s in students_in_group]

    social_data = (
        db.query(SocialPassport, SocialStatus, Person)
        .join(SocialStatus, SocialPassport.social_status_id == SocialStatus.status_id)
        .join(Person, SocialPassport.student_id == Person.person_id)
        .filter(SocialPassport.student_id.in_(student_ids))
        .all()
    )

    status_students = {}
    for sp, status, person in social_data:
        if status.status_name not in status_students:
            status_students[status.status_name] = []
        full_name = f"{person.surname} {person.name} {person.patronymic or ''}".strip()
        if full_name not in status_students[status.status_name]:
            status_students[status.status_name].append(full_name)

    for status_name in status_students:
        status_students[status_name].sort()

    wb = Workbook()
    ws = wb.active
    ws.title = f"СП {group_name}"

    today = date.today().strftime("%d.%m.%Y")

    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(
        start_color="D6BCFA", end_color="D6BCFA", fill_type="solid"
    )

    ws.merge_cells("A1:B1")
    cell = ws["A1"]
    cell.value = "Социальный паспорт группы"
    cell.font = title_font
    cell.alignment = center_align

    cell = ws["C1"]
    cell.value = group_name
    cell.font = Font(bold=True, size=12)
    cell.alignment = center_align

    ws.merge_cells("A2:B2")
    cell = ws["A2"]
    cell.value = "Дата составления"
    cell.font = header_font
    cell.alignment = center_align

    cell = ws["C2"]
    cell.value = today
    cell.alignment = center_align

    statuses = sorted(status_students.keys())
    max_students = max(
        [len(students) for students in status_students.values()], default=0
    )

    for col, status_name in enumerate(statuses, start=1):
        cell = ws.cell(row=4, column=col, value=status_name)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill

    for row in range(max_students):
        for col, status_name in enumerate(statuses, start=1):
            students_list = status_students[status_name]
            value = students_list[row] if row < len(students_list) else ""
            cell = ws.cell(row=5 + row, column=col, value=value)
            cell.alignment = left_align
            cell.border = thin_border

    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output, group_name


def generate_dormitory_list(group_id: int, db: Session) -> tuple:
    group = db.query(Group).filter_by(group_id=group_id).first()
    if not group:
        return None

    group_name = format_group_name(group)

    try:
        current_course = int(group.course.course_name)
    except (ValueError, AttributeError):
        current_course = 1

    dorm_records = (
        db.query(StudentInDormitory, Person)
        .join(Student, StudentInDormitory.student_id == Student.person_id)
        .join(StudentInGroup, Student.person_id == StudentInGroup.student_id)
        .join(Person, Student.person_id == Person.person_id)
        .filter(StudentInGroup.group_id == group_id, Student.is_expelled == False)
        .all()
    )

    students_data = []
    for dorm, person in dorm_records:
        fio = f"{person.surname} {person.name} {person.patronymic or ''}".strip()
        students_data.append({"fio": fio, "room": dorm.room_number})

    students_data.sort(key=lambda x: x["fio"])

    wb = Workbook()
    ws = wb.active
    ws.title = f"СО {group_name}"
    today = date.today().strftime("%d.%m.%Y")

    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(
        start_color="D6BCFA", end_color="D6BCFA", fill_type="solid"
    )

    ws.merge_cells("A1:B1")
    ws["A1"] = "Список студентов в общежитии"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    ws["C1"] = group_name
    ws["C1"].font = Font(bold=True, size=12)
    ws["C1"].alignment = center_align

    ws.merge_cells("A2:B2")
    ws["A2"] = "Дата составления"
    ws["A2"].font = header_font
    ws["A2"].alignment = center_align

    ws["C2"] = today
    ws["C2"].alignment = center_align

    course_cols = ["1 курс", "2 курс", "3 курс", "4 курс"]
    ws.cell(row=4, column=1, value="ФИО").font = header_font
    ws.cell(row=4, column=1).alignment = center_align
    ws.cell(row=4, column=1).border = thin_border
    ws.cell(row=4, column=1).fill = header_fill

    for i, col_name in enumerate(course_cols, start=2):
        cell = ws.cell(row=4, column=i, value=col_name)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill

    for row_idx, student in enumerate(students_data, start=5):
        ws.cell(row=row_idx, column=1, value=student["fio"]).alignment = left_align
        ws.cell(row=row_idx, column=1).border = thin_border

        end_col = current_course + 1
        end_col = min(end_col, 5)

        for col_idx in range(2, end_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=student["room"])
            cell.alignment = center_align
            cell.border = thin_border

        for c in range(end_col + 1, 6):
            ws.cell(row=row_idx, column=c).border = thin_border

    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, group_name


def generate_activists_group(group_id: int, db: Session) -> tuple:
    group = db.query(Group).filter_by(group_id=group_id).first()
    if not group:
        return None

    group_name = format_group_name(group)

    activists = (
        db.query(PostInGroup, Person, PostInGroupType)
        .join(Person, PostInGroup.student_id == Person.person_id)
        .join(
            PostInGroupType,
            PostInGroup.post_in_group_type_id == PostInGroupType.post_in_group_type_id,
        )
        .filter(
            PostInGroup.student_id.in_(
                db.query(StudentInGroup.student_id).filter_by(group_id=group_id)
            )
        )
        .all()
    )

    positions_data = {}
    for activist, person, post_type in activists:
        position_name = post_type.post_in_group_type_name
        course = activist.course_id

        if position_name not in positions_data:
            positions_data[position_name] = {}

        full_name = f"{person.surname} {person.name} {person.patronymic or ''}".strip()
        positions_data[position_name][course] = full_name

    wb = Workbook()
    ws = wb.active
    ws.title = f"АГ {group_name}"
    today = date.today().strftime("%d.%m.%Y")

    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(
        start_color="D6BCFA", end_color="D6BCFA", fill_type="solid"
    )

    ws.merge_cells("A1:B1")
    ws["A1"] = "Активисты группы"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    ws["C1"] = group_name
    ws["C1"].font = Font(bold=True, size=12)
    ws["C1"].alignment = center_align

    ws.merge_cells("A2:B2")
    ws["A2"] = "Дата составления"
    ws["A2"].font = header_font
    ws["A2"].alignment = center_align

    ws["C2"] = today
    ws["C2"].alignment = center_align

    course_cols = ["1 курс", "2 курс", "3 курс", "4 курс"]
    ws.cell(row=4, column=1, value="Социальная нагрузка").font = header_font
    ws.cell(row=4, column=1).alignment = center_align
    ws.cell(row=4, column=1).border = thin_border
    ws.cell(row=4, column=1).fill = header_fill

    for i, col_name in enumerate(course_cols, start=2):
        cell = ws.cell(row=4, column=i, value=col_name)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill

    sorted_positions = sorted(positions_data.keys())
    for row_idx, position in enumerate(sorted_positions, start=5):
        ws.cell(row=row_idx, column=1, value=position).alignment = left_align
        ws.cell(row=row_idx, column=1).border = thin_border

        for course in range(1, 5):
            value = positions_data[position].get(course, "")
            cell = ws.cell(row=row_idx, column=course + 1, value=value)
            cell.alignment = left_align
            cell.border = thin_border

    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, group_name


def generate_extracurricular(group_id: int, db: Session) -> tuple:
    group = db.query(Group).filter_by(group_id=group_id).first()
    if not group:
        return None

    group_name = format_group_name(group)

    students_in_group = db.query(StudentInGroup).filter_by(group_id=group_id).all()
    student_ids = [s.student_id for s in students_in_group]

    hobbies_data = (
        db.query(StudentHobby, HobbyType, Person)
        .join(HobbyType, StudentHobby.hobby_type_id == HobbyType.hobby_type_id)
        .join(Person, StudentHobby.student_id == Person.person_id)
        .filter(StudentHobby.student_id.in_(student_ids))
        .all()
    )

    student_hobbies = {}
    for sh, hobby, person in hobbies_data:
        full_name = f"{person.surname} {person.name} {person.patronymic or ''}".strip()
        if full_name not in student_hobbies:
            student_hobbies[full_name] = []
        student_hobbies[full_name].append(hobby.hobby_type_name)

    for name in student_hobbies:
        student_hobbies[name].sort()

    all_students = (
        db.query(StudentInGroup, Person)
        .join(Person, StudentInGroup.student_id == Person.person_id)
        .join(Student, StudentInGroup.student_id == Student.person_id)
        .filter(StudentInGroup.group_id == group_id, Student.is_expelled == False)
        .all()
    )

    for sig, person in all_students:
        full_name = f"{person.surname} {person.name} {person.patronymic or ''}".strip()
        if full_name not in student_hobbies:
            student_hobbies[full_name] = []

    sorted_students = sorted(student_hobbies.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = f"ВЗ {group_name}"
    today = date.today().strftime("%d.%m.%Y")

    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(
        start_color="D6BCFA", end_color="D6BCFA", fill_type="solid"
    )

    ws.merge_cells("A1:B1")
    ws["A1"] = "Внеучебная занятость"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    ws["C1"] = group_name
    ws["C1"].font = Font(bold=True, size=12)
    ws["C1"].alignment = center_align

    ws.merge_cells("A2:B2")
    ws["A2"] = "Дата составления"
    ws["A2"].font = header_font
    ws["A2"].alignment = center_align

    ws["C2"] = today
    ws["C2"].alignment = center_align

    ws.cell(row=4, column=1, value="ФИО").font = header_font
    ws.cell(row=4, column=1).alignment = center_align
    ws.cell(row=4, column=1).border = thin_border
    ws.cell(row=4, column=1).fill = header_fill

    ws.cell(row=4, column=2, value="Занятость").font = header_font
    ws.cell(row=4, column=2).alignment = center_align
    ws.cell(row=4, column=2).border = thin_border
    ws.cell(row=4, column=2).fill = header_fill

    for row_idx, student_name in enumerate(sorted_students, start=5):
        ws.cell(row=row_idx, column=1, value=student_name).alignment = left_align
        ws.cell(row=row_idx, column=1).border = thin_border

        hobbies_str = ", ".join(student_hobbies[student_name])
        cell = ws.cell(row=row_idx, column=2, value=hobbies_str)
        cell.alignment = left_align
        cell.border = thin_border

    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, group_name


def generate_class_hours_attendance(group_id: int, db: Session) -> tuple:
    group = db.query(Group).filter_by(group_id=group_id).first()
    if not group:
        return None

    group_name = format_group_name(group)

    class_hours = (
        db.query(ClassHour)
        .filter_by(group_id=group_id)
        .order_by(ClassHour.class_hour_date)
        .all()
    )

    students_in_group = (
        db.query(StudentInGroup, Person)
        .join(Person, StudentInGroup.student_id == Person.person_id)
        .join(Student, StudentInGroup.student_id == Student.person_id)
        .filter(StudentInGroup.group_id == group_id, Student.is_expelled == False)
        .all()
    )

    attendance_data = {}
    for sig, person in students_in_group:
        full_name = f"{person.surname} {person.name} {person.patronymic or ''}".strip()
        attendance_data[full_name] = {}

        for ch in class_hours:
            visit = (
                db.query(VisitingClassHour)
                .filter_by(class_hour_id=ch.class_hour_id, student_id=person.person_id)
                .first()
            )

            if visit:
                attendance_data[full_name][ch.class_hour_date] = (
                    "я" if visit.is_visited else "н"
                )
            else:
                attendance_data[full_name][ch.class_hour_date] = ""

    sorted_students = sorted(attendance_data.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = f"УП {group_name}"
    today = date.today().strftime("%d.%m.%Y")

    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(
        start_color="D6BCFA", end_color="D6BCFA", fill_type="solid"
    )

    ws.merge_cells("A1:B1")
    ws["A1"] = "Учёт посещений классных часов"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    ws["C1"] = group_name
    ws["C1"].font = Font(bold=True, size=12)
    ws["C1"].alignment = center_align

    ws.merge_cells("A2:B2")
    ws["A2"] = "Дата составления"
    ws["A2"].font = header_font
    ws["A2"].alignment = center_align

    ws["C2"] = today
    ws["C2"].alignment = center_align

    ws.cell(row=4, column=1, value="ФИО").font = header_font
    ws.cell(row=4, column=1).alignment = center_align
    ws.cell(row=4, column=1).border = thin_border
    ws.cell(row=4, column=1).fill = header_fill

    for col_idx, ch in enumerate(class_hours, start=2):
        date_str = ch.class_hour_date.strftime("%d.%m.%Y") if ch.class_hour_date else ""
        cell = ws.cell(row=4, column=col_idx, value=date_str)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill

    for row_idx, student_name in enumerate(sorted_students, start=5):
        ws.cell(row=row_idx, column=1, value=student_name).alignment = left_align
        ws.cell(row=row_idx, column=1).border = thin_border

        for col_idx, ch in enumerate(class_hours, start=2):
            value = attendance_data[student_name].get(ch.class_hour_date, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_align
            cell.border = thin_border

    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, group_name


def generate_observation_sheet(group_id: int, db: Session) -> tuple:
    group = db.query(Group).filter_by(group_id=group_id).first()
    if not group:
        return None

    group_name = format_group_name(group)

    students_in_group = (
        db.query(StudentInGroup, Person)
        .join(Person, StudentInGroup.student_id == Person.person_id)
        .join(Student, StudentInGroup.student_id == Student.person_id)
        .filter(StudentInGroup.group_id == group_id, Student.is_expelled == False)
        .all()
    )

    student_ids = [sig.student_id for sig, _ in students_in_group]
    student_map = {
        sig.student_id: f"{person.surname} {person.name} {person.patronymic or ''}".strip()
        for sig, person in students_in_group
    }
    student_chars = {sid: "" for sid in student_ids}

    if student_ids:
        obs_records = (
            db.query(ObservationList)
            .filter(ObservationList.student_id.in_(student_ids))
            .order_by(ObservationList.observation_date.desc())
            .all()
        )

        for obs in obs_records:
            if obs.student_id in student_chars and student_chars[obs.student_id] == "":
                student_chars[obs.student_id] = obs.characteristic or ""

    student_data = {student_map[sid]: char for sid, char in student_chars.items()}
    sorted_students = sorted(student_data.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = f"ЛН {group_name}"
    today = date.today().strftime("%d.%m.%Y")

    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(
        start_color="D6BCFA", end_color="D6BCFA", fill_type="solid"
    )

    ws.merge_cells("A1:B1")
    ws["A1"] = "Лист наблюдений куратора"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    ws["C1"] = group_name
    ws["C1"].font = Font(bold=True, size=12)
    ws["C1"].alignment = center_align

    ws.merge_cells("A2:B2")
    ws["A2"] = "Дата составления"
    ws["A2"].font = header_font
    ws["A2"].alignment = center_align

    ws["C2"] = today
    ws["C2"].alignment = center_align

    ws.cell(row=4, column=1, value="ФИО").font = header_font
    ws.cell(row=4, column=1).alignment = center_align
    ws.cell(row=4, column=1).border = thin_border
    ws.cell(row=4, column=1).fill = header_fill

    ws.cell(row=4, column=2, value="Характеристика").font = header_font
    ws.cell(row=4, column=2).alignment = center_align
    ws.cell(row=4, column=2).border = thin_border
    ws.cell(row=4, column=2).fill = header_fill

    for row_idx, student_name in enumerate(sorted_students, start=5):
        ws.cell(row=row_idx, column=1, value=student_name).alignment = left_align
        ws.cell(row=row_idx, column=1).border = thin_border

        char_value = student_data[student_name]
        cell = ws.cell(row=row_idx, column=2, value=char_value)
        cell.alignment = left_align
        cell.border = thin_border

    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, group_name


def generate_parent_meetings(group_id: int, db: Session) -> tuple:
    group = db.query(Group).filter_by(group_id=group_id).first()
    if not group:
        return None

    group_name = format_group_name(group)
    meetings = (
        db.query(ParentMeeting)
        .filter_by(group_id=group_id)
        .order_by(ParentMeeting.meeting_date)
        .all()
    )

    wb = Workbook()
    if meetings:
        wb.remove(wb.active)

    months_ru = [
        "января",
        "февраля",
        "марта",
        "апреля",
        "мая",
        "июня",
        "июля",
        "августа",
        "сентября",
        "октября",
        "ноября",
        "декабря",
    ]
    label_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for meeting in meetings:
        ws = wb.create_sheet()
        d = meeting.meeting_date
        date_str = d.strftime("%d.%m.%Y") if d else "__.__.____"
        day = str(d.day) if d else "__"
        month = months_ru[d.month - 1] if d else "________"
        year = str(d.year) if d else "____"

        ws.title = f"РС {date_str} {group_name}"[:31]

        def set_cell(r, c, val, font=None, align=None, merge=None):
            cell = ws.cell(row=r, column=c, value=val)
            if font:
                cell.font = font
            if align:
                cell.alignment = align
            cell.border = thin_border
            if merge:
                ws.merge_cells(merge)
            return cell

        set_cell(
            1, 1, "Протокол родительского собрания №", label_font, align_center, "A1:B1"
        )
        set_cell(1, 3, str(meeting.parent_meeting_id), align=align_center)
        set_cell(2, 1, f'От "{day}" {month} {year}', align=align_center, merge="A2:C2")
        set_cell(3, 1, "", merge="A3:C3")

        set_cell(4, 1, "Приглашённые", label_font, align_left)
        set_cell(4, 2, meeting.invited or "", align=align_left, merge="B4:C4")

        set_cell(5, 1, "Посетило", label_font, align_center)
        set_cell(5, 2, meeting.visited_count or 0, align=align_center)
        set_cell(5, 3, "")

        set_cell(6, 1, "Отсутствовали", label_font, align_center)
        set_cell(6, 2, meeting.unvisited or 0, align=align_center)
        set_cell(6, 3, "")

        set_cell(7, 1, "Отсутствовали по уважительной причине", label_font, align_left)
        set_cell(7, 2, meeting.excused_count or 0, align=align_center, merge="B7:C7")

        set_cell(8, 1, "", merge="A8:C8")
        set_cell(9, 1, "Тема собрания", label_font, align_left)
        set_cell(9, 2, meeting.topics or "", align=align_left, merge="B9:C9")

        set_cell(10, 1, "По теме собрания выступили", label_font, align_left)
        set_cell(10, 2, meeting.speakers or "", align=align_left, merge="B10:C10")

        for r in range(11, 14):
            set_cell(r, 1, "", merge=f"A{r}:C{r}")

        set_cell(14, 1, "В ходе собрания решено", label_font, align_left)
        set_cell(14, 2, meeting.meeting_result or "", align=align_left, merge="B14:C14")
        set_cell(15, 1, "", merge="A15:C15")

        curator = db.query(Curator).filter_by(person_id=group.curator_id).first()
        curator_fio = ""
        if curator:
            p = db.query(Person).filter_by(person_id=curator.person_id).first()
            if p:
                curator_fio = f"{p.surname} {p.name} {p.patronymic or ''}".strip()

        set_cell(16, 1, "Куратор", label_font, align_left)
        set_cell(16, 2, curator_fio, align=align_left)
        set_cell(16, 3, "______", align=align_center)

        set_cell(17, 1, "Председатель родительского комитета", label_font, align_left)
        set_cell(17, 2, "______", align=align_center)
        set_cell(17, 3, "______", align=align_center)

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 22
        ws.row_dimensions[10].height = 70
        ws.row_dimensions[14].height = 70

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, group_name


def generate_individual_work(group_id: int, db: Session) -> tuple:
    group = db.query(Group).filter_by(group_id=group_id).first()
    if not group:
        return None

    group_name = format_group_name(group)
    today = date.today().strftime("%d.%m.%Y")

    irs_records = (
        db.query(StudentIndividualWork, Person)
        .join(Student, StudentIndividualWork.student_id == Student.person_id)
        .join(StudentInGroup, Student.person_id == StudentInGroup.student_id)
        .join(Person, Student.person_id == Person.person_id)
        .filter(StudentInGroup.group_id == group_id)
        .order_by(StudentIndividualWork.date.desc())
        .all()
    )

    irr_records = (
        db.query(ParentIndividualWork, Parent)
        .join(Parent, ParentIndividualWork.parent_id == Parent.parent_id)
        .join(StudentParent, Parent.parent_id == StudentParent.parent_id)
        .join(StudentInGroup, StudentParent.student_id == StudentInGroup.student_id)
        .filter(StudentInGroup.group_id == group_id)
        .order_by(ParentIndividualWork.date.desc())
        .all()
    )

    wb = Workbook()

    def fill_sheet(ws, sheet_title, header_title, records):
        ws.title = sheet_title[:31]
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        header_fill = PatternFill(
            start_color="D6BCFA", end_color="D6BCFA", fill_type="solid"
        )

        ws.merge_cells("A1:B1")
        ws["A1"] = header_title
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align

        ws["C1"] = group_name
        ws["C1"].font = Font(bold=True, size=12)
        ws["C1"].alignment = center_align

        ws.merge_cells("A2:B2")
        ws["A2"] = "Дата составления"
        ws["A2"].font = header_font
        ws["A2"].alignment = center_align

        ws["C2"] = today
        ws["C2"].alignment = center_align

        headers = ["Дата", "ФИО", "Тема", "Решение"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = header_fill

        for row_idx, (record, person_or_parent) in enumerate(records, start=5):
            date_val = record.date.strftime("%d.%m.%Y") if record.date else ""
            fio = f"{person_or_parent.surname} {person_or_parent.name} {person_or_parent.patronymic or ''}".strip()
            topic = record.topic or ""
            result = record.result or ""

            ws.cell(row=row_idx, column=1, value=date_val).alignment = center_align
            ws.cell(row=row_idx, column=1).border = thin_border
            ws.cell(row=row_idx, column=2, value=fio).alignment = left_align
            ws.cell(row=row_idx, column=2).border = thin_border
            ws.cell(row=row_idx, column=3, value=topic).alignment = left_align
            ws.cell(row=row_idx, column=3).border = thin_border
            ws.cell(row=row_idx, column=4, value=result).alignment = left_align
            ws.cell(row=row_idx, column=4).border = thin_border

        for col_idx in range(1, 5):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    fill_sheet(
        wb.active,
        f"ИРС {group_name}",
        "Индивидуальная работа со студентами",
        irs_records,
    )
    ws2 = wb.create_sheet()
    fill_sheet(
        ws2, f"ИРР {group_name}", "Индивидуальная работа с родителями", irr_records
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, group_name


def generate_general_report(group_id: int, db: Session) -> tuple:
    group = db.query(Group).filter_by(group_id=group_id).first()
    if not group:
        return None

    group_name = format_group_name(group)
    master_wb = Workbook()

    def copy_sheet(source_bytesio, master_wb):
        source_wb = load_workbook(source_bytesio)
        for sheet_name in source_wb.sheetnames:
            source_ws = source_wb[sheet_name]
            safe_title = sheet_name[:31]
            target_ws = master_wb.create_sheet(title=safe_title)

            for row in source_ws:
                for cell in row:
                    target_cell = target_ws[cell.coordinate]
                    target_cell.value = cell.value
                    if cell.has_style:
                        target_cell.font = copy(cell.font)
                        target_cell.border = copy(cell.border)
                        target_cell.fill = copy(cell.fill)
                        target_cell.number_format = cell.number_format
                        target_cell.protection = copy(cell.protection)
                        target_cell.alignment = copy(cell.alignment)

            for col_name, col_dim in source_ws.column_dimensions.items():
                target_ws.column_dimensions[col_name].width = col_dim.width
            for row_idx, row_dim in source_ws.row_dimensions.items():
                target_ws.row_dimensions[row_idx].height = row_dim.height

            for merge_range in source_ws.merged_cells.ranges:
                target_ws.merge_cells(str(merge_range))

    sp_data, _ = generate_social_passport(group_id, db)
    copy_sheet(sp_data, master_wb)

    dorm_data, _ = generate_dormitory_list(group_id, db)
    copy_sheet(dorm_data, master_wb)

    act_data, _ = generate_activists_group(group_id, db)
    copy_sheet(act_data, master_wb)

    extra_data, _ = generate_extracurricular(group_id, db)
    copy_sheet(extra_data, master_wb)

    hours_data, _ = generate_class_hours_attendance(group_id, db)
    copy_sheet(hours_data, master_wb)

    obs_data, _ = generate_observation_sheet(group_id, db)
    copy_sheet(obs_data, master_wb)

    iw_data, _ = generate_individual_work(group_id, db)
    copy_sheet(iw_data, master_wb)

    pm_data, _ = generate_parent_meetings(group_id, db)
    copy_sheet(pm_data, master_wb)

    if "Sheet" in master_wb.sheetnames:
        del master_wb["Sheet"]

    output = io.BytesIO()
    master_wb.save(output)
    output.seek(0)

    return output, group_name
