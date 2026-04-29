import io
from sqlalchemy.orm import Session
from openpyxl import Workbook
from database.models import ParentMeeting, Curator, Person
from .helpers import get_group_info
from . import styles

def generate_parent_meetings(group_id: int, db: Session) -> tuple[io.BytesIO, str] | None:
    group, group_name = get_group_info(group_id, db)
    if not group:
        return None

    meetings = db.query(ParentMeeting).filter_by(group_id=group_id).order_by(ParentMeeting.meeting_date).all()
    if not meetings:
        return None

    wb = Workbook()
    wb.remove(wb.active)

    months_ru = ["января", "февраля", "марта", "апреля", "мая", "июня", "июля", "августа", "сентября", "октября", "ноября", "декабря"]

    for meeting in meetings:
        ws = wb.create_sheet()
        d = meeting.meeting_date
        date_str = d.strftime("%d.%m.%Y") if d else "__.__.____"
        ws.title = f"РС {date_str} {group_name}"[:31]

        def set_cell(r, c, val, font=None, align=None, merge=None):
            cell = ws.cell(row=r, column=c, value=val)
            if font: cell.font = font
            if align: cell.alignment = align
            cell.border = styles.THIN_BORDER
            if merge: ws.merge_cells(merge)
            return cell

        set_cell(1, 1, "Протокол родительского собрания №", styles.HEADER_FONT, styles.CENTER_ALIGN, "A1:B1")
        set_cell(1, 3, str(meeting.parent_meeting_id), align=styles.CENTER_ALIGN)
        set_cell(2, 1, f'От "{d.day if d else "__"} {months_ru[d.month-1] if d else "________"} {d.year if d else "____"}"', align=styles.CENTER_ALIGN, merge="A2:C2")
        set_cell(3, 1, "", merge="A3:C3")

        set_cell(4, 1, "Приглашённые", styles.HEADER_FONT, styles.TOP_LEFT_ALIGN)
        set_cell(4, 2, meeting.invited or "", align=styles.TOP_LEFT_ALIGN, merge="B4:C4")
        set_cell(5, 1, "Посетило", styles.HEADER_FONT, styles.CENTER_ALIGN)
        set_cell(5, 2, meeting.visited_count or 0, align=styles.CENTER_ALIGN)
        set_cell(5, 3, "")
        set_cell(6, 1, "Отсутствовали", styles.HEADER_FONT, styles.CENTER_ALIGN)
        set_cell(6, 2, meeting.unvisited or 0, align=styles.CENTER_ALIGN)
        set_cell(6, 3, "")
        set_cell(7, 1, "Отсутствовали по уважительной причине", styles.HEADER_FONT, styles.TOP_LEFT_ALIGN)
        set_cell(7, 2, meeting.excused_count or 0, align=styles.CENTER_ALIGN, merge="B7:C7")
        set_cell(8, 1, "", merge="A8:C8")
        set_cell(9, 1, "Тема собрания", styles.HEADER_FONT, styles.TOP_LEFT_ALIGN)
        set_cell(9, 2, meeting.topics or "", align=styles.TOP_LEFT_ALIGN, merge="B9:C9")
        set_cell(10, 1, "По теме собрания выступили", styles.HEADER_FONT, styles.TOP_LEFT_ALIGN)
        set_cell(10, 2, meeting.speakers or "", align=styles.TOP_LEFT_ALIGN, merge="B10:C10")

        for r in range(11, 14): set_cell(r, 1, "", merge=f"A{r}:C{r}")

        set_cell(14, 1, "В ходе собрания решено", styles.HEADER_FONT, styles.TOP_LEFT_ALIGN)
        set_cell(14, 2, meeting.meeting_result or "", align=styles.TOP_LEFT_ALIGN, merge="B14:C14")
        set_cell(15, 1, "", merge="A15:C15")

        curator = db.query(Curator).filter_by(person_id=group.curator_id).first()
        curator_fio = ""
        if curator:
            p = db.query(Person).filter_by(person_id=curator.person_id).first()
            if p: curator_fio = f"{p.surname} {p.name} {p.patronymic or ''}".strip()

        set_cell(16, 1, "Куратор", styles.HEADER_FONT, styles.TOP_LEFT_ALIGN)
        set_cell(16, 2, curator_fio, align=styles.TOP_LEFT_ALIGN)
        set_cell(16, 3, "______", align=styles.CENTER_ALIGN)
        set_cell(17, 1, "Председатель родительского комитета", styles.HEADER_FONT, styles.TOP_LEFT_ALIGN)
        set_cell(17, 2, "______", align=styles.CENTER_ALIGN)
        set_cell(17, 3, "______", align=styles.CENTER_ALIGN)

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 22
        ws.row_dimensions[10].height = 70
        ws.row_dimensions[14].height = 70

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, group_name