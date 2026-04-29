import io
from sqlalchemy.orm import Session
from openpyxl import Workbook
from database.models import ClassHour, VisitingClassHour, StudentInGroup, Person, Student
from .helpers import get_group_info, setup_standard_header, format_full_name, auto_adjust_columns, apply_cell_style
from . import styles

def generate_class_hours_attendance(group_id: int, db: Session) -> tuple[io.BytesIO, str] | None:
    group, group_name = get_group_info(group_id, db)
    if not group:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = f"УП {group_name}"
    setup_standard_header(ws, "Учёт посещений классных часов", group_name)

    class_hours = db.query(ClassHour).filter_by(group_id=group_id).order_by(ClassHour.class_hour_date).all()
    students_in_group = (
        db.query(StudentInGroup, Person)
        .join(Person, StudentInGroup.student_id == Person.person_id)
        .join(Student, StudentInGroup.student_id == Student.person_id)
        .filter(StudentInGroup.group_id == group_id, Student.is_expelled == False)
        .all()
    )

    attendance_data: dict[str, dict] = {}
    for sig, person in students_in_group:
        full_name = format_full_name(person.surname, person.name, person.patronymic)
        attendance_data[full_name] = {}
        for ch in class_hours:
            visit = db.query(VisitingClassHour).filter_by(class_hour_id=ch.class_hour_id, student_id=person.person_id).first()
            attendance_data[full_name][ch.class_hour_date] = "я" if visit and visit.is_visited else ("н" if visit else "")

    apply_cell_style(ws.cell(row=4, column=1), "ФИО", styles.CENTER_ALIGN)
    ws.cell(row=4, column=1).fill = styles.HEADER_FILL

    for col_idx, ch in enumerate(class_hours, start=2):
        date_str = ch.class_hour_date.strftime("%d.%m.%Y") if ch.class_hour_date else ""
        apply_cell_style(ws.cell(row=4, column=col_idx), date_str, styles.CENTER_ALIGN)
        ws.cell(row=4, column=col_idx).fill = styles.HEADER_FILL

    for row_idx, student_name in enumerate(sorted(attendance_data.keys()), start=5):
        apply_cell_style(ws.cell(row=row_idx, column=1), student_name)
        for col_idx, ch in enumerate(class_hours, start=2):
            apply_cell_style(ws.cell(row=row_idx, column=col_idx), attendance_data[student_name].get(ch.class_hour_date, ""), styles.CENTER_ALIGN)

    auto_adjust_columns(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, group_name