import io
from sqlalchemy.orm import Session
from openpyxl import Workbook
from database.models import StudentInGroup, SocialPassport, SocialStatus, Person
from .helpers import get_group_info, setup_standard_header, format_full_name, auto_adjust_columns, apply_cell_style
from . import styles

def generate_social_passport(group_id: int, db: Session) -> tuple[io.BytesIO, str] | None:
    group, group_name = get_group_info(group_id, db)
    if not group:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = f"СП {group_name}"
    setup_standard_header(ws, "Социальный паспорт группы", group_name)

    student_ids = [s.student_id for s in db.query(StudentInGroup).filter_by(group_id=group_id).all()]
    social_data = (
        db.query(SocialPassport, SocialStatus, Person)
        .join(SocialStatus, SocialPassport.social_status_id == SocialStatus.status_id)
        .join(Person, SocialPassport.student_id == Person.person_id)
        .filter(SocialPassport.student_id.in_(student_ids))
        .all()
    )

    status_students: dict[str, list[str]] = {}
    for _, status, person in social_data:
        full_name = format_full_name(person.surname, person.name, person.patronymic)
        status_students.setdefault(status.status_name, []).append(full_name)

    for lst in status_students.values():
        lst.sort()

    statuses = sorted(status_students.keys())
    max_students = max((len(s) for s in status_students.values()), default=0)

    for col, status_name in enumerate(statuses, start=1):
        apply_cell_style(ws.cell(row=4, column=col), status_name, styles.CENTER_ALIGN)
        ws.cell(row=4, column=col).fill = styles.HEADER_FILL

    for row in range(max_students):
        for col, status_name in enumerate(statuses, start=1):
            value = status_students[status_name][row] if row < len(status_students[status_name]) else ""
            apply_cell_style(ws.cell(row=5 + row, column=col), value)

    auto_adjust_columns(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, group_name