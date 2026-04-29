import io
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from database.models import StudentIndividualWork, Student, StudentInGroup, Person, ParentIndividualWork, Parent, StudentParent
from .helpers import get_group_info, setup_standard_header, format_full_name, auto_adjust_columns, apply_cell_style
from . import styles

def _fill_sheet(ws: Worksheet, sheet_title: str, header_title: str, records, group_name: str, today: str) -> None:
    ws.title = sheet_title[:31]
    setup_standard_header(ws, header_title, group_name)
    ws["C2"] = today

    headers = ["Дата", "ФИО", "Тема", "Решение"]
    for col_idx, header in enumerate(headers, start=1):
        apply_cell_style(ws.cell(row=4, column=col_idx), header, styles.CENTER_ALIGN)
        ws.cell(row=4, column=col_idx).fill = styles.HEADER_FILL

    for row_idx, (record, person_or_parent) in enumerate(records, start=5):
        date_val = record.date.strftime("%d.%m.%Y") if record.date else ""
        fio = format_full_name(person_or_parent.surname, person_or_parent.name, person_or_parent.patronymic)
        apply_cell_style(ws.cell(row=row_idx, column=1), date_val, styles.CENTER_ALIGN)
        apply_cell_style(ws.cell(row=row_idx, column=2), fio)
        apply_cell_style(ws.cell(row=row_idx, column=3), record.topic or "")
        apply_cell_style(ws.cell(row=row_idx, column=4), record.result or "")

    auto_adjust_columns(ws)

def generate_individual_work(group_id: int, db: Session) -> tuple[io.BytesIO, str] | None:
    group, group_name = get_group_info(group_id, db)
    if not group:
        return None

    today = io.datetime.date.today().strftime("%d.%m.%Y")
    wb = Workbook()

    irs_records = (
        db.query(StudentIndividualWork, Person)
        .join(Student, StudentIndividualWork.student_id == Student.person_id)
        .join(StudentInGroup, Student.person_id == StudentInGroup.student_id)
        .join(Person, Student.person_id == Person.person_id)
        .filter(StudentInGroup.group_id == group_id)
        .order_by(StudentIndividualWork.date.desc())
        .all()
    )
    irr_records = (
        db.query(ParentIndividualWork, Parent)
        .join(Parent, ParentIndividualWork.parent_id == Parent.parent_id)
        .join(StudentParent, Parent.parent_id == StudentParent.parent_id)
        .join(StudentInGroup, StudentParent.student_id == StudentInGroup.student_id)
        .filter(StudentInGroup.group_id == group_id)
        .order_by(ParentIndividualWork.date.desc())
        .all()
    )

    _fill_sheet(wb.active, f"ИРС {group_name}", "Индивидуальная работа со студентами", irs_records, group_name, today)
    ws2 = wb.create_sheet()
    _fill_sheet(ws2, f"ИРР {group_name}", "Индивидуальная работа с родителями", irr_records, group_name, today)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, group_name