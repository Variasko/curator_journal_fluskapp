import io
from sqlalchemy.orm import Session
from openpyxl import Workbook
from database.models import StudentInGroup, Person, Student, ObservationList
from .helpers import get_group_info, setup_standard_header, format_full_name, auto_adjust_columns, apply_cell_style
from . import styles

def generate_observation_sheet(group_id: int, db: Session) -> tuple[io.BytesIO, str] | None:
    group, group_name = get_group_info(group_id, db)
    if not group:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = f"ЛН {group_name}"
    setup_standard_header(ws, "Лист наблюдений куратора", group_name)

    students_in_group = (
        db.query(StudentInGroup, Person)
        .join(Person, StudentInGroup.student_id == Person.person_id)
        .join(Student, StudentInGroup.student_id == Student.person_id)
        .filter(StudentInGroup.group_id == group_id, Student.is_expelled == False)
        .all()
    )

    student_ids = [sig.student_id for sig, _ in students_in_group]
    student_map = {sig.student_id: format_full_name(p.surname, p.name, p.patronymic) for sig, p in students_in_group}
    student_chars = {sid: "" for sid in student_ids}

    if student_ids:
        for obs in db.query(ObservationList).filter(ObservationList.student_id.in_(student_ids)).order_by(ObservationList.observation_date.desc()).all():
            if obs.student_id in student_chars and not student_chars[obs.student_id]:
                student_chars[obs.student_id] = obs.characteristic or ""

    student_data = {student_map[sid]: char for sid, char in student_chars.items()}

    apply_cell_style(ws.cell(row=4, column=1), "ФИО", styles.CENTER_ALIGN)
    ws.cell(row=4, column=1).fill = styles.HEADER_FILL
    apply_cell_style(ws.cell(row=4, column=2), "Характеристика", styles.CENTER_ALIGN)
    ws.cell(row=4, column=2).fill = styles.HEADER_FILL

    for row_idx, student_name in enumerate(sorted(student_data.keys()), start=5):
        apply_cell_style(ws.cell(row=row_idx, column=1), student_name)
        apply_cell_style(ws.cell(row=row_idx, column=2), student_data[student_name])

    auto_adjust_columns(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, group_name