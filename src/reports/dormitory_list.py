import io
from sqlalchemy.orm import Session
from openpyxl import Workbook
from database.models import StudentInDormitory, Student, StudentInGroup, Person
from .helpers import get_group_info, setup_standard_header, format_full_name, auto_adjust_columns, apply_cell_style
from . import styles

def generate_dormitory_list(group_id: int, db: Session) -> tuple[io.BytesIO, str] | None:
    group, group_name = get_group_info(group_id, db)
    if not group:
        return None

    try:
        current_course = int(group.course.course_name)
    except (ValueError, AttributeError):
        current_course = 1

    wb = Workbook()
    ws = wb.active
    ws.title = f"СО {group_name}"
    setup_standard_header(ws, "Список студентов в общежитии", group_name)

    dorm_records = (
        db.query(StudentInDormitory, Person)
        .join(Student, StudentInDormitory.student_id == Student.person_id)
        .join(StudentInGroup, Student.person_id == StudentInGroup.student_id)
        .join(Person, Student.person_id == Person.person_id)
        .filter(StudentInGroup.group_id == group_id, Student.is_expelled == False)
        .all()
    )

    students_data = [{"fio": format_full_name(p.surname, p.name, p.patronymic), "room": d.room_number}
                     for d, p in dorm_records]
    students_data.sort(key=lambda x: x["fio"])

    course_cols = ["1 курс", "2 курс", "3 курс", "4 курс"]
    apply_cell_style(ws.cell(row=4, column=1), "ФИО", styles.CENTER_ALIGN)
    ws.cell(row=4, column=1).fill = styles.HEADER_FILL

    for i, col_name in enumerate(course_cols, start=2):
        apply_cell_style(ws.cell(row=4, column=i), col_name, styles.CENTER_ALIGN)
        ws.cell(row=4, column=i).fill = styles.HEADER_FILL

    for row_idx, student in enumerate(students_data, start=5):
        apply_cell_style(ws.cell(row=row_idx, column=1), student["fio"])
        end_col = min(current_course + 1, 5)
        for col_idx in range(2, end_col + 1):
            apply_cell_style(ws.cell(row=row_idx, column=col_idx), student["room"], styles.CENTER_ALIGN)
        for c in range(end_col + 1, 6):
            ws.cell(row=row_idx, column=c).border = styles.THIN_BORDER

    auto_adjust_columns(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, group_name