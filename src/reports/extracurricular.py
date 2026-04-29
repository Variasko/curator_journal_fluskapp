import io
from sqlalchemy.orm import Session
from openpyxl import Workbook
from database.models import StudentInGroup, StudentHobby, HobbyType, Person, Student
from .helpers import get_group_info, setup_standard_header, format_full_name, auto_adjust_columns, apply_cell_style
from . import styles

def generate_extracurricular(group_id: int, db: Session) -> tuple[io.BytesIO, str] | None:
    group, group_name = get_group_info(group_id, db)
    if not group:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = f"ВЗ {group_name}"
    setup_standard_header(ws, "Внеучебная занятость", group_name)

    student_ids = [s.student_id for s in db.query(StudentInGroup).filter_by(group_id=group_id).all()]
    hobbies_data = (
        db.query(StudentHobby, HobbyType, Person)
        .join(HobbyType, StudentHobby.hobby_type_id == HobbyType.hobby_type_id)
        .join(Person, StudentHobby.student_id == Person.person_id)
        .filter(StudentHobby.student_id.in_(student_ids))
        .all()
    )

    student_hobbies: dict[str, list[str]] = {}
    for _, hobby, person in hobbies_data:
        full_name = format_full_name(person.surname, person.name, person.patronymic)
        student_hobbies.setdefault(full_name, []).append(hobby.hobby_type_name)

    all_students = (
        db.query(StudentInGroup, Person)
        .join(Person, StudentInGroup.student_id == Person.person_id)
        .join(Student, StudentInGroup.student_id == Student.person_id)
        .filter(StudentInGroup.group_id == group_id, Student.is_expelled == False)
        .all()
    )
    for _, person in all_students:
        full_name = format_full_name(person.surname, person.name, person.patronymic)
        student_hobbies.setdefault(full_name, [])

    apply_cell_style(ws.cell(row=4, column=1), "ФИО", styles.CENTER_ALIGN)
    ws.cell(row=4, column=1).fill = styles.HEADER_FILL
    apply_cell_style(ws.cell(row=4, column=2), "Занятость", styles.CENTER_ALIGN)
    ws.cell(row=4, column=2).fill = styles.HEADER_FILL

    for row_idx, student_name in enumerate(sorted(student_hobbies.keys()), start=5):
        apply_cell_style(ws.cell(row=row_idx, column=1), student_name)
        apply_cell_style(ws.cell(row=row_idx, column=2), ", ".join(sorted(student_hobbies[student_name])))

    auto_adjust_columns(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, group_name