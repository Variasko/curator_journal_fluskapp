import io
from datetime import date
from copy import copy
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook import Workbook

from database.models import Group
from utils.formatters import format_group_name

from . import styles

def get_group_info(group_id: int, db: Session) -> tuple[Group | None, str | None]:
    group = db.query(Group).filter_by(group_id=group_id).first()
    if not group:
        return None, None
    return group, format_group_name(group)

def format_full_name(surname: str, name: str, patronymic: str | None) -> str:
    return f"{surname} {name} {patronymic or ''}".strip()

def setup_standard_header(ws: Worksheet, title: str, group_name: str) -> None:
    today = date.today().strftime("%d.%m.%Y")
    ws.merge_cells("A1:B1")
    ws["A1"] = title
    ws["A1"].font = styles.TITLE_FONT
    ws["A1"].alignment = styles.CENTER_ALIGN

    ws["C1"] = group_name
    ws["C1"].font = styles.GROUP_FONT
    ws["C1"].alignment = styles.CENTER_ALIGN

    ws.merge_cells("A2:B2")
    ws["A2"] = "Дата составления"
    ws["A2"].font = styles.HEADER_FONT
    ws["A2"].alignment = styles.CENTER_ALIGN

    ws["C2"] = today
    ws["C2"].alignment = styles.CENTER_ALIGN

def apply_cell_style(cell, value: str | None, alignment=styles.LEFT_ALIGN, border=styles.THIN_BORDER):
    cell.value = value
    cell.alignment = alignment
    cell.border = border

def auto_adjust_columns(ws: Worksheet, max_width: int = 50) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                try:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, max_width)

def copy_worksheet_to_master(source_bytesio: io.BytesIO, master_wb: Workbook) -> None:
    source_wb = load_workbook(source_bytesio)
    for sheet_name in source_wb.sheetnames:
        source_ws = source_wb[sheet_name]
        target_ws = master_wb.create_sheet(title=sheet_name[:31])

        for row in source_ws:
            for cell in row:
                target_cell = target_ws[cell.coordinate]
                target_cell.value = cell.value
                if cell.has_style:
                    target_cell.font = copy(cell.font)
                    target_cell.border = copy(cell.border)
                    target_cell.fill = copy(cell.fill)
                    target_cell.number_format = cell.number_format
                    target_cell.protection = copy(cell.protection)
                    target_cell.alignment = copy(cell.alignment)

        for col_name, col_dim in source_ws.column_dimensions.items():
            target_ws.column_dimensions[col_name].width = col_dim.width
        for row_idx, row_dim in source_ws.row_dimensions.items():
            target_ws.row_dimensions[row_idx].height = row_dim.height

        for merge_range in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merge_range))