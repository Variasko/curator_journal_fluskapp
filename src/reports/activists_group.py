import io
from sqlalchemy.orm import Session
from openpyxl import Workbook
from database.models import PostInGroup, Person, PostInGroupType, StudentInGroup
from .helpers import get_group_info, setup_standard_header, format_full_name, auto_adjust_columns, apply_cell_style
from . import styles

def generate_activists_group(group_id: int, db: Session) -> tuple[io.BytesIO, str] | None:
    group, group_name = get_group_info(group_id, db)
    if not group:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = f"АГ {group_name}"
    setup_standard_header(ws, "Активисты группы", group_name)

    activists = (
        db.query(PostInGroup, Person, PostInGroupType)
        .join(Person, PostInGroup.student_id == Person.person_id)
        .join(PostInGroupType, PostInGroup.post_in_group_type_id == PostInGroupType.post_in_group_type_id)
        .filter(PostInGroup.student_id.in_(
            db.query(StudentInGroup.student_id).filter_by(group_id=group_id)
        ))
        .all()
    )

    positions_data: dict[str, dict[int, str]] = {}
    for activist, person, post_type in activists:
        full_name = format_full_name(person.surname, person.name, person.patronymic)
        positions_data.setdefault(post_type.post_in_group_type_name, {})[activist.course_id] = full_name

    course_cols = ["1 курс", "2 курс", "3 курс", "4 курс"]
    apply_cell_style(ws.cell(row=4, column=1), "Социальная нагрузка", styles.CENTER_ALIGN)
    ws.cell(row=4, column=1).fill = styles.HEADER_FILL
    for i, col_name in enumerate(course_cols, start=2):
        apply_cell_style(ws.cell(row=4, column=i), col_name, styles.CENTER_ALIGN)
        ws.cell(row=4, column=i).fill = styles.HEADER_FILL

    sorted_positions = sorted(positions_data.keys())
    for row_idx, position in enumerate(sorted_positions, start=5):
        apply_cell_style(ws.cell(row=row_idx, column=1), position)
        for course in range(1, 5):
            apply_cell_style(ws.cell(row=row_idx, column=course + 1), positions_data[position].get(course, ""))

    auto_adjust_columns(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, group_name